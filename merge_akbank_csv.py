#!/usr/bin/env python3
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Read both CSV files
daily_totals = defaultdict(float)

# File 1: 801808b0-AKBANK_POS.csv
print("Reading AKBANK_POS.csv...")
with open('/root/.claude/uploads/d3c97f87-f4b1-50a0-88f9-174947e8589b/801808b0-AKBANK_POS.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # Skip header
    for row in reader:
        if len(row) >= 2 and row[0].strip():
            tarih = row[0].strip()
            tutar_str = row[1].strip()
            # Remove TL suffix and convert
            tutar_str = tutar_str.replace(' TL', '').strip()
            tutar_str = tutar_str.replace('.', '').replace(',', '.')
            try:
                tutar = float(tutar_str)
                # Extract day from "1.07.2026" format
                day = int(tarih.split('.')[0])
                daily_totals[day] += tutar
            except:
                pass

# File 2: 566cf84d-AKBANK_POS_2.csv
print("Reading AKBANK_POS_2.csv...")
with open('/root/.claude/uploads/d3c97f87-f4b1-50a0-88f9-174947e8589b/566cf84d-AKBANK_POS_2.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # Skip header
    for row in reader:
        if len(row) >= 2 and row[0].strip():
            tarih = row[0].strip()
            tutar_str = row[1].strip()
            # Remove TL suffix and convert
            tutar_str = tutar_str.replace(' TL', '').strip()
            tutar_str = tutar_str.replace('.', '').replace(',', '.')
            try:
                tutar = float(tutar_str)
                # Extract day from "1.07.2026" format
                day = int(tarih.split('.')[0])
                daily_totals[day] += tutar
            except:
                pass

# Group by date ranges
date_ranges = {
    "01-10.07": [],
    "11-20.07": [],
    "21-31.07": [],
}

for day in sorted(daily_totals.keys()):
    tutar = daily_totals[day]
    if 1 <= day <= 10:
        date_ranges["01-10.07"].append((day, tutar))
    elif 11 <= day <= 20:
        date_ranges["11-20.07"].append((day, tutar))
    else:
        date_ranges["21-31.07"].append((day, tutar))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Günlük Brut Tutarlar"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
range_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
range_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
total_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18

row = 1

# Calculate grand total
grand_total = sum(daily_totals.values())

# Add monthly total header
ws[f'A{row}'] = "AYLIK TOPLAM"
ws[f'B{row}'] = grand_total
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
row += 2

# Add headers
ws[f'A{row}'] = "Tarih"
ws[f'B{row}'] = "Brüt Tutar"
for col in ['A', 'B']:
    ws[f'{col}{row}'].font = header_font
    ws[f'{col}{row}'].fill = header_fill
    ws[f'{col}{row}'].border = border
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

# Add data for each date range
total_sum = 0
for range_name in ["01-10.07", "11-20.07", "21-31.07"]:
    # Add range header
    ws[f'A{row}'] = range_name
    ws[f'A{row}'].font = range_font
    ws[f'A{row}'].fill = range_fill
    ws[f'A{row}'].border = border
    row += 1

    # Add daily data
    range_sum = 0
    for day, tutar in date_ranges[range_name]:
        ws[f'A{row}'] = f"{day:02d}/07/2026"
        ws[f'B{row}'] = tutar
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0.00 "₺"'
        range_sum += tutar
        total_sum += tutar
        row += 1

    # Add range subtotal
    ws[f'A{row}'] = f"{range_name} TOPLAM"
    ws[f'B{row}'] = range_sum
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws[f'B{row}'].font = Font(bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'B{row}'].number_format = '#,##0.00 "₺"'
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 2

# Add grand total
ws[f'A{row}'] = "GENEL TOPLAM"
ws[f'B{row}'] = total_sum
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'A{row}'].border = border
ws[f'B{row}'].border = border
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right')

# Save workbook
wb.save('/home/user/-mery-lmaz/ISIK_PETROL-AKBANK_POS-Gunluk_Brut-Temmuz2026.xlsx')

print("Excel file created successfully!")
print(f"\n=== VERIFICATION ===")
print(f"AKBANK 1 subtotal: {sum([t for d,t in [(d, daily_totals[d]) for d in sorted(daily_totals.keys())] if d <= 10]):.2f} ₺")
print(f"AKBANK 2 subtotal: {sum([t for d,t in [(d, daily_totals[d]) for d in sorted(daily_totals.keys())] if d > 10]):.2f} ₺")
print(f"Merged total: {total_sum:,.2f} ₺")
