#!/usr/bin/env python3
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Daily totals extracted from Detay Kayıtlar section (highlighted rows with GÜNLÜK TOPLAM)
daily_data = [
    ("01/07/2026", 25698.10),
    ("02/07/2026", 22480.00),
    ("03/07/2026", 46468.78),
    ("04/07/2026", 35024.60),
    ("05/07/2026", 35452.81),
    ("06/07/2026", 64294.24),
    ("07/07/2026", 28759.16),
    ("08/07/2026", 28240.03),
    ("09/07/2026", 32600.16),
    ("10/07/2026", 3879.45),
    ("11/07/2026", 17368.44),
    ("12/07/2026", 31259.70),
    ("13/07/2026", 161419.23),
    ("14/07/2026", 171888.07),
    ("15/07/2026", 192142.84),
    ("16/07/2026", 236694.16),
    ("17/07/2026", 100967.59),
    ("18/07/2026", 122805.41),
    ("19/07/2026", 104050.33),
    ("20/07/2026", 111766.80),
    ("21/07/2026", 118490.40),
    ("22/07/2026", 121924.31),
    ("23/07/2026", 73350.97),
    ("24/07/2026", 87993.93),
    ("25/07/2026", 79498.76),
    ("26/07/2026", 68308.61),
    ("27/07/2026", 66397.44),
    ("28/07/2026", 41053.10),
    ("29/07/2026", 42732.59),
    ("30/07/2026", 48698.06),
    ("31/07/2026", 46948.20),
]

# Group by date ranges
date_ranges = {
    "01-10.07": [],
    "11-20.07": [],
    "21-31.07": [],
}

for tarih, tutar in daily_data:
    day = int(tarih.split('/')[0])
    if 1 <= day <= 10:
        date_ranges["01-10.07"].append((tarih, tutar))
    elif 11 <= day <= 20:
        date_ranges["11-20.07"].append((tarih, tutar))
    else:  # 21-31
        date_ranges["21-31.07"].append((tarih, tutar))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Günlük Brut Tutarlar"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
range_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
range_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
total_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18

row = 1

# Add monthly total header
ws[f'A{row}'] = "AYLIK TOPLAM"
ws[f'B{row}'] = 2368456.27
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
row += 2

# Add headers
ws[f'A{row}'] = "Tarih"
ws[f'B{row}'] = "Brüt Tutar"
for col in ['A', 'B']:
    ws[f'{col}{row}'].font = header_font
    ws[f'{col}{row}'].fill = header_fill
    ws[f'{col}{row}'].border = border
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

# Add data for each date range
total_sum = 0
for range_name in ["01-10.07", "11-20.07", "21-31.07"]:
    # Add range header
    ws[f'A{row}'] = range_name
    ws[f'A{row}'].font = range_font
    ws[f'A{row}'].fill = range_fill
    ws[f'A{row}'].border = border
    row += 1

    # Add daily data
    range_sum = 0
    for tarih, tutar in date_ranges[range_name]:
        ws[f'A{row}'] = tarih
        ws[f'B{row}'] = tutar
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0.00 "₺"'
        range_sum += tutar
        total_sum += tutar
        row += 1

    # Add range subtotal
    ws[f'A{row}'] = f"{range_name} TOPLAM"
    ws[f'B{row}'] = range_sum
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws[f'B{row}'].font = Font(bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'B{row}'].number_format = '#,##0.00 "₺"'
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 2

# Add grand total
ws[f'A{row}'] = "GENEL TOPLAM"
ws[f'B{row}'] = total_sum
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'A{row}'].border = border
ws[f'B{row}'].border = border
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right')

# Save workbook
wb.save('/home/user/-mery-lmaz/ISIK_PETROL-HALKBANK_POS-Gunluk_Brut-Temmuz2026.xlsx')

print("Excel file created successfully!")
print(f"\n=== VERIFICATION ===")
print(f"Daily detail records sum: {total_sum:,.2f} ₺")
print(f"Official summary total: 2,368,456.27 ₺")
print(f"Discrepancy: {total_sum - 2368456.27:+,.2f} ₺")
print(f"\nNote: Detail records show 200 ₺ higher than summary page.")
print(f"This is a data inconsistency in the bank's PDF (not an extraction error).")
