#!/usr/bin/env python3
"""
Muhasebe TR - Excel Export Modülü
Makaleler ve S&C verilerini Excel dosyasına aktarır
"""

import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self):
        self.data_dir = Path("/home/user/-mery-lmaz/muhasebe_data")
        self.data_dir.mkdir(exist_ok=True)

    def create_styled_worksheet(self, wb, title):
        """Stil uygulanmış worksheet oluştur"""
        ws = wb.create_sheet(title)

        # Font stili
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        return ws, header_font, header_fill, thin_border

    def export_articles_to_excel(self, articles, filename="muhasebe_makaleler.xlsx"):
        """Makaleleri Excel'e aktar"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Varsayılan sheet'i kaldır

            # Tüm makaleler sheet'i
            ws, header_font, header_fill, thin_border = self.create_styled_worksheet(
                wb, "Tüm Makaleler"
            )

            # Başlıkları ekle
            headers = ["Sıra", "Başlık", "Açıklama", "Tarih", "Kategori", "Kaynak", "URL"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Veri ekle
            for row, article in enumerate(articles, 2):
                ws.cell(row=row, column=1).value = row - 1
                ws.cell(row=row, column=2).value = article.get('title', '')
                ws.cell(row=row, column=3).value = article.get('description', '')
                ws.cell(row=row, column=4).value = article.get('date', '')
                ws.cell(row=row, column=5).value = article.get('category', 'Diğer')
                ws.cell(row=row, column=6).value = article.get('source', '')
                ws.cell(row=row, column=7).value = article.get('url', '')

                # Border ve alignment ekle
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Kolon genişliklerini ayarla
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 30

            # Kategoriye göre sheet'ler
            by_category = {}
            for article in articles:
                cat = article.get('category', 'Diğer')
                if cat not in by_category:
                    by_category[cat] = []
                by_category[cat].append(article)

            for category in sorted(by_category.keys()):
                ws, header_font, header_fill, thin_border = self.create_styled_worksheet(
                    wb, category[:31]  # Excel sheet adı 31 karakter sınırı
                )

                headers = ["Sıra", "Başlık", "Açıklama", "Tarih", "URL"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

                for row, article in enumerate(by_category[category], 2):
                    ws.cell(row=row, column=1).value = row - 1
                    ws.cell(row=row, column=2).value = article.get('title', '')
                    ws.cell(row=row, column=3).value = article.get('description', '')
                    ws.cell(row=row, column=4).value = article.get('date', '')
                    ws.cell(row=row, column=5).value = article.get('url', '')

                    for col in range(1, 6):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 40
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 30

            # İstatistikler sheet'i
            ws, header_font, header_fill, thin_border = self.create_styled_worksheet(
                wb, "İstatistikler"
            )

            ws.cell(row=1, column=1).value = "Kategori"
            ws.cell(row=1, column=2).value = "Makale Sayısı"

            for col in range(1, 3):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row, (category, articles_in_cat) in enumerate(sorted(by_category.items()), 2):
                ws.cell(row=row, column=1).value = category
                ws.cell(row=row, column=2).value = len(articles_in_cat)

                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            ws.cell(row=len(by_category) + 3, column=1).value = "TOPLAM"
            ws.cell(row=len(by_category) + 3, column=2).value = len(articles)

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15

            # Excel dosyasını kaydet
            output_path = self.data_dir / filename
            wb.save(output_path)
            print(f"✅ Excel dosyası kaydedildi: {output_path}")
            return output_path

        except Exception as e:
            print(f"❌ Excel export hatası: {e}")
            return None

    def export_qa_to_excel(self, qa_data, filename="muhasebe_sorular.xlsx"):
        """S&C verilerini Excel'e aktar"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tüm Sorular"

            # Başlıkları ekle
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = ["Sıra", "Soru", "Cevap", "Tarih", "Saat", "Email"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # S&C verilerini ekle
            questions = qa_data.get("questions", [])
            for row, qa in enumerate(questions, 2):
                ws.cell(row=row, column=1).value = qa.get('id', row - 1)
                ws.cell(row=row, column=2).value = qa.get('question', '')
                ws.cell(row=row, column=3).value = qa.get('answer', '')

                # Tarih ve saati ayır
                timestamp = qa.get('timestamp', '')
                if timestamp:
                    date_part = timestamp[:10]
                    time_part = timestamp[11:16]
                else:
                    date_part = ''
                    time_part = ''

                ws.cell(row=row, column=4).value = date_part
                ws.cell(row=row, column=5).value = time_part
                ws.cell(row=row, column=6).value = qa.get('user_email', '')

                # Border ve alignment ekle
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Kolon genişliklerini ayarla
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 20

            # Excel dosyasını kaydet
            output_path = self.data_dir / filename
            wb.save(output_path)
            print(f"✅ S&C Excel dosyası kaydedildi: {output_path}")
            return output_path

        except Exception as e:
            print(f"❌ S&C Excel export hatası: {e}")
            return None


# Test
if __name__ == "__main__":
    print("\n" + "="*60)
    print("MUHASEBE TR - EXCEL EXPORT MODÜLÜ")
    print("="*60 + "\n")

    exporter = ExcelExporter()

    # Test makaleler
    test_articles = [
        {
            'title': 'Asgari Ücret Güncellemesi 2026',
            'description': 'Yeni asgari ücret oranları açıklandı',
            'date': '18.08.2026',
            'url': 'https://www.muhasebetr.com/asgari-ucret/',
            'source': 'muhasebetr.com',
            'category': 'Asgari Ücret'
        },
        {
            'title': 'KDV Oranlarında Değişiklikler',
            'description': 'KDV oranlarında yeni düzenlemeler',
            'date': '17.08.2026',
            'url': 'https://www.muhasebetr.com/kdv/',
            'source': 'muhasebetr.com',
            'category': 'Vergi'
        },
        {
            'title': 'Demirbaş Amortisman',
            'description': '2026 amortisman oranları',
            'date': '16.08.2026',
            'url': 'https://www.muhasebetr.com/amortisman/',
            'source': 'muhasebetr.com',
            'category': 'Amortisman'
        }
    ]

    print("📊 Makaleler Excel'e aktarılıyor...\n")
    exporter.export_articles_to_excel(test_articles)

    # Test S&C verisi
    test_qa = {
        "questions": [
            {
                "id": 1,
                "question": "Asgari ücret ne kadar?",
                "answer": "2026 yılı asgari ücret...",
                "timestamp": "2026-08-18T10:30:00",
                "user_email": "user@example.com"
            },
            {
                "id": 2,
                "question": "KDV oranı kaçtır?",
                "answer": "KDV oranı %18'dir...",
                "timestamp": "2026-08-18T11:00:00",
                "user_email": "user2@example.com"
            }
        ]
    }

    print("📊 Sorular-Cevaplar Excel'e aktarılıyor...\n")
    exporter.export_qa_to_excel(test_qa)

    print("\n✅ Excel export işlemi tamamlandı!")
