#!/usr/bin/env python3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load both AKBANK files
wb1 = load_workbook('ISIK_PETROL_AKBANK_OZET.xlsx')
wb2 = load_workbook('ISIK_PETROL_AKBANK2_OZET.xlsx')

ws1 = wb1.active
ws2 = wb2.active

# Extract data from both sheets
data1 = {
    "01.06.2026 - 10.06.2026": (136, 27199.67),
    "11.06.2026 - 20.06.2026": (102, 10495.88),
    "21.06.2026 - 30.06.2026": (86, 7013.45),
}

data2 = {
    "01.06.2026 - 10.06.2026": (152, 16215.74),
    "11.06.2026 - 20.06.2026": (93, 17467.22),
    "21.06.2026 - 30.06.2026": (96, 18168.03),
}

# Merge data
merged = {}
for date_range in data1.keys():
    işlem1, tutar1 = data1[date_range]
    işlem2, tutar2 = data2[date_range]
    merged[date_range] = (işlem1 + işlem2, tutar1 + tutar2)

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Özet"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
total_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18

# Header
ws['A1'] = 'Tarih Aralığı'
ws['B1'] = 'İşlem Sayısı'
ws['C1'] = 'Toplam Tutar (₺)'

for col in ['A', 'B', 'C']:
    ws[f'{col}1'].font = header_font
    ws[f'{col}1'].fill = header_fill
    ws[f'{col}1'].border = border
    ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

# Data rows
row = 2
total_işlem = 0
total_tutar = 0

for date_range in ["01.06.2026 - 10.06.2026", "11.06.2026 - 20.06.2026", "21.06.2026 - 30.06.2026"]:
    işlem, tutar = merged[date_range]

    ws[f'A{row}'] = date_range
    ws[f'B{row}'] = işlem
    ws[f'C{row}'] = tutar

    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'C{row}'].border = border
    ws[f'B{row}'].alignment = Alignment(horizontal='center')
    ws[f'C{row}'].number_format = '#,##0.00'

    total_işlem += işlem
    total_tutar += tutar
    row += 1

# Empty row
row += 1

# Total row
ws[f'A{row}'] = 'GENEL TOPLAM'
ws[f'B{row}'] = total_işlem
ws[f'C{row}'] = total_tutar

for col in ['A', 'B', 'C']:
    ws[f'{col}{row}'].font = total_font
    ws[f'{col}{row}'].fill = total_fill
    ws[f'{col}{row}'].border = border
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

ws[f'C{row}'].number_format = '#,##0.00'

# Save
wb.save('ISIK_PETROL-AKBANK_MERGED-Haziran2026.xlsx')

print("AKBANK dosyaları birleştirildi!")
print(f"\nÖzet:")
print(f"01.06-10.06: {merged['01.06.2026 - 10.06.2026'][0]} işlem, {merged['01.06.2026 - 10.06.2026'][1]:,.2f} ₺")
print(f"11.06-20.06: {merged['11.06.2026 - 20.06.2026'][0]} işlem, {merged['11.06.2026 - 20.06.2026'][1]:,.2f} ₺")
print(f"21.06-30.06: {merged['21.06.2026 - 30.06.2026'][0]} işlem, {merged['21.06.2026 - 30.06.2026'][1]:,.2f} ₺")
print(f"\nGENEL TOPLAM: {total_işlem} işlem, {total_tutar:,.2f} ₺")
