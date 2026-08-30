#!/usr/bin/env python3
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load ZIRAAT file
print("Reading ZIRAAT BANKASI Tutar Komisyon data...")
wb_ziraat = load_workbook('/root/.claude/uploads/d3c97f87-f4b1-50a0-88f9-174947e8589b/81567351-___k_petrol.xlsx')
ws = wb_ziraat.active

# Extract commission rows (only Tutar rows, not POS SATIŞ)
daily_totals = {}
pattern = r'Komisyon\s*:\s*([0-9]+[,\.][0-9]+)'

for row in ws.iter_rows(min_row=2, values_only=False):
    tarih_cell = row[0]
    aciklama_cell = row[1]

    if tarih_cell.value and aciklama_cell.value:
        # Parse date
        date_obj = tarih_cell.value
        if isinstance(date_obj, str):
            date_obj = datetime.fromisoformat(date_obj.replace('Z', '+00:00'))

        day = date_obj.day
        aciklama = str(aciklama_cell.value).strip()

        # Only count rows with "Tutar" (not POS SATIŞ)
        if 'POS SATIŞ' in aciklama:
            continue

        # Look for Komisyon pattern
        match = re.search(pattern, aciklama)
        if match:
            amount_str = match.group(1)
            # Parse Turkish number format (comma is decimal separator)
            amount_str = amount_str.replace(',', '.')
            try:
                amount = float(amount_str)
                if day not in daily_totals:
                    daily_totals[day] = 0
                daily_totals[day] += amount
            except:
                pass

# Group by date ranges
date_ranges = {
    "01-10.07": [],
    "11-20.07": [],
    "21-31.07": [],
}

for day in sorted(daily_totals.keys()):
    tutar = daily_totals[day]
    if 1 <= day <= 10:
        date_ranges["01-10.07"].append((day, tutar))
    elif 11 <= day <= 20:
        date_ranges["11-20.07"].append((day, tutar))
    else:
        date_ranges["21-31.07"].append((day, tutar))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "KOMİSYON"

# Define styles
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_font = Font(bold=True, size=11)
range_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
range_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
total_font = Font(bold=True, size=10)
grand_total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
grand_total_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18

row = 1

# Calculate grand total
grand_total = sum(daily_totals.values())

# Add monthly total header
ws[f'A{row}'] = "AYLIK TOPLAM"
ws[f'B{row}'] = grand_total
ws[f'A{row}'].font = header_font
ws[f'B{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws[f'B{row}'].fill = header_fill
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
row += 2

# Add headers
ws[f'A{row}'] = "Tarih"
ws[f'B{row}'] = "Komisyon"
for col in ['A', 'B']:
    ws[f'{col}{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'{col}{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'{col}{row}'].border = border
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

# Add data for each date range
total_sum = 0
for range_name in ["01-10.07", "11-20.07", "21-31.07"]:
    # Add range header
    ws[f'A{row}'] = range_name
    ws[f'A{row}'].font = range_font
    ws[f'A{row}'].fill = range_fill
    ws[f'A{row}'].border = border
    row += 1

    # Add daily data
    range_sum = 0
    for day, tutar in date_ranges[range_name]:
        ws[f'A{row}'] = f"{day:02d}/07/2026"
        ws[f'B{row}'] = tutar
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0.00 "₺"'
        range_sum += tutar
        total_sum += tutar
        row += 1

    # Add range subtotal
    ws[f'A{row}'] = f"{range_name} TOPLAM"
    ws[f'B{row}'] = range_sum
    ws[f'A{row}'].font = total_font
    ws[f'B{row}'].font = total_font
    ws[f'A{row}'].fill = total_fill
    ws[f'B{row}'].fill = total_fill
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'B{row}'].number_format = '#,##0.00 "₺"'
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 2

# Add grand total
ws[f'A{row}'] = "GENEL TOPLAM"
ws[f'B{row}'] = total_sum
ws[f'A{row}'].font = grand_total_font
ws[f'B{row}'].font = grand_total_font
ws[f'A{row}'].fill = grand_total_fill
ws[f'B{row}'].fill = grand_total_fill
ws[f'A{row}'].border = border
ws[f'B{row}'].border = border
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right')

# Save workbook
wb.save('/home/user/-mery-lmaz/ZİRAAT-TUTAR-KOMİSYON-Temmuz2026.xlsx')

print("Excel file created successfully!")
print(f"\n=== ÖZET ===")
print(f"01-10.07: {sum([t for d,t in date_ranges['01-10.07']]):,.2f} ₺")
print(f"11-20.07: {sum([t for d,t in date_ranges['11-20.07']]):,.2f} ₺")
print(f"21-31.07: {sum([t for d,t in date_ranges['21-31.07']]):,.2f} ₺")
print(f"GENEL TOPLAM: {total_sum:,.2f} ₺")
