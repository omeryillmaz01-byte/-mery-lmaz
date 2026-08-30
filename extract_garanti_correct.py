#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Daily totals from HESAP EKSTRESİ section - grouping by day
# Extracting all Brüt Tutar values by day
daily_totals = defaultdict(float)

# HESAP EKSTRESİ data (day, tutar)
transactions = [
    (1, 1400.00),
    (2, 650.00),
    (2, 1500.00),
    (2, 1400.00),
    (2, 0.00),
    (3, 2150.00),
    (4, 2000.00),
    (4, 1000.00),
    (5, 189.64),
    (5, 810.36),
    (6, 3806.45),
    (6, 200.00),
    (7, 2000.00),
    (8, 3700.00),
    (8, 2140.00),
    (8, 201.05),
    (8, 2668.95),
    (9, 1500.00),
    (9, 750.00),
    (9, 1400.00),
    (10, 5300.00),
    (10, 1150.00),
    (11, 3139.86),
    (12, 1500.00),
    (12, 1112.08),
    (13, 1500.00),
    (13, 1310.43),
    (13, 1400.00),
    (14, 425.00),
    (14, 1275.00),
    (15, 1550.00),
    (15, 450.00),
    (15, 1400.00),
    (16, 2700.00),
    (16, 2132.39),
    (16, 1300.00),
    (17, 3350.00),
    (17, 803.32),
    (19, 246.49),
    (20, 3360.00),
    (20, 2030.00),
    (21, 900.00),
    (21, 539.81),
    (21, 2200.00),
    (22, 7871.85),
    (22, 2548.00),
    (23, 2526.61),
    (23, 4823.39),
    (24, 6038.69),
    (24, 1661.31),
    (25, 3796.37),
    (25, 2829.83),
    (25, 7260.88),
    (26, 200.00),
    (26, 6559.41),
    (27, 150.00),
    (27, 2150.00),
    (27, 553.34),
    (28, 1500.00),
    (28, 3662.08),
    (28, 1000.00),
    (29, 5903.34),
    (29, 3073.76),
    (29, 200.00),
    (30, 4991.81),
    (30, 2058.19),
    (30, 560.94),
    (31, 4629.75),
    (31, 860.00),
    (31, 240.00),
]

# Sum by day
for day, tutar in transactions:
    daily_totals[day] += tutar

# Convert to list and sort
daily_list = sorted([(day, tutar) for day, tutar in daily_totals.items()])

# Group by date ranges
date_ranges = {
    "01-10.07": [],
    "11-20.07": [],
    "21-31.07": [],
}

for day, tutar in daily_list:
    if 1 <= day <= 10:
        date_ranges["01-10.07"].append((day, tutar))
    elif 11 <= day <= 20:
        date_ranges["11-20.07"].append((day, tutar))
    else:
        date_ranges["21-31.07"].append((day, tutar))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Günlük Brut Tutarlar"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
range_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
range_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
total_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18

row = 1

# Add monthly total header
ws[f'A{row}'] = "AYLIK TOPLAM"
ws[f'B{row}'] = 148190.38
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
row += 2

# Add headers
ws[f'A{row}'] = "Tarih"
ws[f'B{row}'] = "Brüt Tutar"
for col in ['A', 'B']:
    ws[f'{col}{row}'].font = header_font
    ws[f'{col}{row}'].fill = header_fill
    ws[f'{col}{row}'].border = border
    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
row += 1

# Add data for each date range
total_sum = 0
for range_name in ["01-10.07", "11-20.07", "21-31.07"]:
    # Add range header
    ws[f'A{row}'] = range_name
    ws[f'A{row}'].font = range_font
    ws[f'A{row}'].fill = range_fill
    ws[f'A{row}'].border = border
    row += 1

    # Add daily data
    range_sum = 0
    for day, tutar in date_ranges[range_name]:
        ws[f'A{row}'] = f"{day:02d}/07/2026"
        ws[f'B{row}'] = tutar
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0.00 "₺"'
        range_sum += tutar
        total_sum += tutar
        row += 1

    # Add range subtotal
    ws[f'A{row}'] = f"{range_name} TOPLAM"
    ws[f'B{row}'] = range_sum
    ws[f'A{row}'].font = Font(bold=True, size=10)
    ws[f'B{row}'].font = Font(bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    ws[f'B{row}'].number_format = '#,##0.00 "₺"'
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 2

# Add grand total
ws[f'A{row}'] = "GENEL TOPLAM"
ws[f'B{row}'] = total_sum
ws[f'A{row}'].font = total_font
ws[f'B{row}'].font = total_font
ws[f'A{row}'].fill = total_fill
ws[f'B{row}'].fill = total_fill
ws[f'A{row}'].border = border
ws[f'B{row}'].border = border
ws[f'B{row}'].number_format = '#,##0.00 "₺"'
ws[f'A{row}'].alignment = Alignment(horizontal='center')
ws[f'B{row}'].alignment = Alignment(horizontal='right')

# Save workbook
wb.save('/home/user/-mery-lmaz/ISIK_PETROL-GARANTI_POS-Gunluk_Brut-Temmuz2026.xlsx')

print("Excel file created successfully!")
print(f"\n=== VERIFICATION ===")
print(f"Extracted total: {total_sum:,.2f} ₺")
print(f"Expected total: 148,190.38 ₺")
print(f"Match: {'✓ EXACT MATCH' if abs(total_sum - 148190.38) < 0.01 else '✗ MISMATCH'}")
